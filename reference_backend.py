from __future__ import annotations

import argparse
import csv
import io
import json
import random
import re
import threading
import time
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bib_ref_checker import (
    iter_bib_entries,
    BibEntry,
    CheckResult,
    ReferenceSearcher,
    decide,
    normalize_title,
    parse_bibtex,
    reference_category,
    record_kind,
)


VERDICT_PRIORITY = {
    "FOUND_MATCH": 3,
    "FOUND_MISMATCH": 2,
    "NOT_FOUND": 1,
}
SUPPORTED_SOURCES = {"arxiv", "serpapi", "scholar-html", "crossref", "openalex"}


@dataclass
class DuplicateGroup:
    duplicate_type: str
    value: str
    keys: List[str]
    titles: List[str]


@dataclass
class SourceOption:
    name: str
    enabled: bool = True
    api_key: str = ""
    proxy: str = ""
    extra: Dict[str, Any] = field(default_factory=dict)


@dataclass
class AnalysisConfig:
    source_order: List[SourceOption]
    delay: float = 2.0
    jitter: float = 1.0
    timeout: float = 15.0
    retries: int = 0
    retry_backoff: float = 0.5
    num_results: int = 8
    year_window: int = 1
    title_threshold: float = 0.92
    loose_title_threshold: float = 0.84
    venue_threshold: float = 0.72
    min_author_overlap: int = 1
    author_check_count: int = 3
    no_author_in_query: bool = False
    user_agent: str = ""
    skip_url_check: bool = False
    cache_path: str = ".api_refcheck_cache.json"
    include_missing_title: bool = False
    verbose: bool = False


@dataclass
class AnalysisSummary:
    total_entries: int
    unique_entries: int
    duplicate_groups: int
    duplicate_entries: int
    verdict_counts: Dict[str, int]


@dataclass
class AnalysisOutput:
    summary: AnalysisSummary
    duplicates: List[DuplicateGroup]
    entries: List[Dict[str, Any]]

    def to_json_bytes(self) -> bytes:
        payload = {
            "summary": asdict(self.summary),
            "duplicates": [asdict(x) for x in self.duplicates],
            "entries": self.entries,
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    def to_csv_bytes(self) -> bytes:
        if not self.entries:
            return b""
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(self.entries[0].keys()))
        writer.writeheader()
        writer.writerows(self.entries)
        return buf.getvalue().encode("utf-8-sig")

    def to_xlsx_bytes(self) -> bytes:
        workbook = Workbook()
        verdict_groups = [
            ("FOUND_MATCH", "完全匹配", "完全匹配", "EAF7EE"),
            ("FOUND_MISMATCH", "部分匹配", "部分匹配", "FFF7E6"),
            ("NOT_FOUND", "没有找到", "没有找到", "FDECEC"),
        ]
        columns = ["编号", "标题", "作者", "年份", "期刊", "原始 .bib", "标记", "搜索源"]

        header_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")
        header_font = Font(bold=True, color="111827")
        wrap_alignment = Alignment(vertical="top", wrap_text=True)

        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for verdict, sheet_name, label, fill_color in verdict_groups:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(columns)
            rows = [row for row in self.entries if row.get("verdict") == verdict]
            verdict_fill = PatternFill(fill_type="solid", fgColor=fill_color)

            for col_idx, title in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = wrap_alignment

            for index, row in enumerate(rows, start=1):
                authors = row.get("input_authors_list") or row.get("input_authors") or ""
                if isinstance(authors, list):
                    authors = ", ".join(authors)
                values = [
                    index,
                    row.get("input_title") or row.get("matched_title") or "",
                    authors,
                    row.get("input_year") or row.get("matched_year") or "",
                    row.get("input_venue") or row.get("matched_venue") or "",
                    row.get("raw_bib") or "",
                    label,
                    row.get("matched_source") or row.get("backend") or "",
                ]
                sheet.append(values)
                current_row = sheet.max_row
                for col_idx in range(1, len(columns) + 1):
                    cell = sheet.cell(row=current_row, column=col_idx)
                    cell.alignment = wrap_alignment
                sheet.cell(row=current_row, column=7).fill = verdict_fill

            widths = {
                1: 10,
                2: 48,
                3: 36,
                4: 10,
                5: 28,
                6: 70,
                7: 14,
                8: 14,
            }
            for col_idx, width in widths.items():
                sheet.column_dimensions[chr(64 + col_idx)].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


@dataclass
class AnalysisJob:
    job_id: str
    filename: str
    status: str
    created_at: float
    updated_at: float
    config: AnalysisConfig
    error: str = ""
    output: Optional[AnalysisOutput] = None
    total_entries: int = 0
    processed_entries: int = 0
    progress: float = 0.0
    progress_message: str = ""
    original_bib_path: str = ""
    modified_bib_path: str = ""


def mask_secret(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 4:
        return "*" * len(value)
    return f"{value[:2]}{'*' * (len(value) - 4)}{value[-2:]}"


def serialize_source_option(source: SourceOption, include_secrets: bool = False) -> Dict[str, Any]:
    return {
        "name": source.name,
        "enabled": source.enabled,
        "api_key": source.api_key if include_secrets else mask_secret(source.api_key),
        "proxy": source.proxy,
        "extra": source.extra,
    }


def serialize_config(config: AnalysisConfig, include_secrets: bool = False) -> Dict[str, Any]:
    return {
        "source_order": [
            serialize_source_option(source, include_secrets=include_secrets)
            for source in config.source_order
        ],
        "delay": config.delay,
        "jitter": config.jitter,
        "timeout": config.timeout,
        "retries": config.retries,
        "retry_backoff": config.retry_backoff,
        "num_results": config.num_results,
        "year_window": config.year_window,
        "title_threshold": config.title_threshold,
        "loose_title_threshold": config.loose_title_threshold,
        "venue_threshold": config.venue_threshold,
        "min_author_overlap": config.min_author_overlap,
        "author_check_count": config.author_check_count,
        "no_author_in_query": config.no_author_in_query,
        "user_agent": config.user_agent,
        "skip_url_check": config.skip_url_check,
        "cache_path": config.cache_path,
        "include_missing_title": config.include_missing_title,
        "verbose": config.verbose,
    }


class JobStore:
    def __init__(self, base_dir: str = ".job_store") -> None:
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self._jobs: Dict[str, AnalysisJob] = {}
        self._lock = threading.Lock()
        self._load_existing_jobs()

    def _job_dir(self, job_id: str) -> Path:
        return self.base_dir / job_id

    def _job_meta_path(self, job_id: str) -> Path:
        return self._job_dir(job_id) / "job.json"

    def _job_payload(self, job: AnalysisJob) -> Dict[str, Any]:
        return {
            "job_id": job.job_id,
            "filename": job.filename,
            "status": job.status,
            "created_at": job.created_at,
            "updated_at": job.updated_at,
            "config": serialize_config(job.config, include_secrets=True),
            "error": job.error,
            "total_entries": job.total_entries,
            "processed_entries": job.processed_entries,
            "progress": job.progress,
            "progress_message": job.progress_message,
            "original_bib_path": job.original_bib_path,
            "modified_bib_path": job.modified_bib_path,
            "output": None
            if job.output is None
            else {
                "summary": asdict(job.output.summary),
                "duplicates": [asdict(x) for x in job.output.duplicates],
                "entries": job.output.entries,
            },
        }

    def _job_from_payload(self, payload: Dict[str, Any]) -> AnalysisJob:
        config = AnalysisConfig(
            source_order=build_source_options(payload["config"]["source_order"]),
            delay=payload["config"]["delay"],
            jitter=payload["config"]["jitter"],
            timeout=payload["config"]["timeout"],
            retries=payload["config"]["retries"],
            retry_backoff=payload["config"]["retry_backoff"],
            num_results=payload["config"]["num_results"],
            year_window=payload["config"]["year_window"],
            title_threshold=payload["config"]["title_threshold"],
            loose_title_threshold=payload["config"]["loose_title_threshold"],
            venue_threshold=payload["config"]["venue_threshold"],
            min_author_overlap=payload["config"]["min_author_overlap"],
            author_check_count=payload["config"]["author_check_count"],
            no_author_in_query=payload["config"]["no_author_in_query"],
            user_agent=payload["config"]["user_agent"],
            skip_url_check=payload["config"]["skip_url_check"],
            cache_path=payload["config"]["cache_path"],
            include_missing_title=payload["config"]["include_missing_title"],
            verbose=payload["config"]["verbose"],
        )
        output = None
        if payload.get("output"):
            output = AnalysisOutput(
                summary=AnalysisSummary(**payload["output"]["summary"]),
                duplicates=[DuplicateGroup(**x) for x in payload["output"]["duplicates"]],
                entries=payload["output"]["entries"],
            )
        return AnalysisJob(
            job_id=payload["job_id"],
            filename=payload["filename"],
            status=payload["status"],
            created_at=payload["created_at"],
            updated_at=payload["updated_at"],
            config=config,
            error=payload.get("error", ""),
            output=output,
            total_entries=payload.get("total_entries", 0),
            processed_entries=payload.get("processed_entries", 0),
            progress=payload.get("progress", 0.0),
            progress_message=payload.get("progress_message", ""),
            original_bib_path=payload.get("original_bib_path", ""),
            modified_bib_path=payload.get("modified_bib_path", ""),
        )

    def _persist(self, job: AnalysisJob) -> None:
        job_dir = self._job_dir(job.job_id)
        job_dir.mkdir(parents=True, exist_ok=True)
        self._job_meta_path(job.job_id).write_text(
            json.dumps(self._job_payload(job), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        if job.output is not None:
            (job_dir / "report.json").write_bytes(job.output.to_json_bytes())
            (job_dir / "report.csv").write_bytes(job.output.to_csv_bytes())

    def _load_existing_jobs(self) -> None:
        for path in self.base_dir.glob("*/job.json"):
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
                job = self._job_from_payload(payload)
                self._jobs[job.job_id] = job
            except Exception:
                continue

    def _original_bib_path(self, job_id: str) -> Path:
        return self._job_dir(job_id) / "original.bib"

    def _modified_bib_path(self, job_id: str) -> Path:
        return self._job_dir(job_id) / "modified.bib"

    def create(self, filename: str, config: AnalysisConfig, original_bib_text: str) -> AnalysisJob:
        now = time.time()
        job_id = uuid.uuid4().hex
        job = AnalysisJob(
            job_id=job_id,
            filename=filename,
            status="pending",
            created_at=now,
            updated_at=now,
            config=config,
            original_bib_path=str(self._original_bib_path(job_id)),
            modified_bib_path=str(self._modified_bib_path(job_id)),
        )
        with self._lock:
            self._jobs[job.job_id] = job
            self._job_dir(job.job_id).mkdir(parents=True, exist_ok=True)
            self._original_bib_path(job.job_id).write_text(original_bib_text, encoding="utf-8")
            self._modified_bib_path(job.job_id).write_text(original_bib_text, encoding="utf-8")
            self._persist(job)
        return job

    def get(self, job_id: str) -> Optional[AnalysisJob]:
        with self._lock:
            return self._jobs.get(job_id)

    def list(self) -> List[AnalysisJob]:
        with self._lock:
            return sorted(
                self._jobs.values(),
                key=lambda job: (job.updated_at, job.created_at),
                reverse=True,
            )

    def update(self, job_id: str, **fields: Any) -> Optional[AnalysisJob]:
        with self._lock:
            job = self._jobs.get(job_id)
            if not job:
                return None
            for key, value in fields.items():
                setattr(job, key, value)
            job.updated_at = time.time()
            self._persist(job)
            return job

    def read_modified_bib(self, job_id: str) -> str:
        job = self.get(job_id)
        if not job:
            raise FileNotFoundError(job_id)
        path = self._ensure_modified_bib(job)
        return path.read_text(encoding="utf-8")

    def read_original_bib(self, job_id: str) -> str:
        job = self.get(job_id)
        if not job or not job.original_bib_path:
            raise FileNotFoundError(job_id)
        return Path(job.original_bib_path).read_text(encoding="utf-8")

    def write_modified_bib(self, job_id: str, text: str) -> None:
        job = self.get(job_id)
        if not job:
            raise FileNotFoundError(job_id)
        path = self._ensure_modified_bib(job)
        path.write_text(text, encoding="utf-8")

    def modified_bib_file(self, job_id: str) -> Path:
        job = self.get(job_id)
        if not job:
            raise FileNotFoundError(job_id)
        return self._ensure_modified_bib(job)

    def _ensure_modified_bib(self, job: AnalysisJob) -> Path:
        if job.modified_bib_path:
            path = Path(job.modified_bib_path)
            if path.exists():
                return path
        path = self._modified_bib_path(job.job_id)
        source_text = ""
        if job.original_bib_path and Path(job.original_bib_path).exists():
            source_text = Path(job.original_bib_path).read_text(encoding="utf-8")
        path.write_text(source_text, encoding="utf-8")
        job.modified_bib_path = str(path)
        self.update(job.job_id, modified_bib_path=str(path))
        return path


def build_source_options(raw_sources: Optional[List[Dict[str, Any]]]) -> List[SourceOption]:
    if not raw_sources:
        return [
            SourceOption(name="arxiv"),
            SourceOption(name="serpapi"),
            SourceOption(name="scholar-html"),
        ]
    options = []
    seen = set()
    for item in raw_sources:
        name = item["name"]
        if name not in SUPPORTED_SOURCES:
            raise ValueError(f"Unsupported source: {name}")
        if name in seen:
            continue
        seen.add(name)
        options.append(
            SourceOption(
                name=name,
                enabled=item.get("enabled", True),
                api_key=item.get("api_key", ""),
                proxy=item.get("proxy", ""),
                extra=item.get("extra", {}),
            )
        )
    if not any(option.enabled for option in options):
        raise ValueError("At least one search source must be enabled.")
    return options


def load_entries_from_bib_text(text: str, include_missing_title: bool = False) -> List[BibEntry]:
    entries = parse_bibtex(text)
    if not include_missing_title:
        entries = [e for e in entries if e.title]
    return entries


def serialize_bib_entry(entry: BibEntry) -> Dict[str, Any]:
    payload = asdict(entry)
    payload["reference_category"] = reference_category(entry)
    return payload


def detect_duplicates(entries: List[BibEntry]) -> List[DuplicateGroup]:
    groups: List[DuplicateGroup] = []
    for duplicate_type, key_fn in (
        ("doi", lambda e: e.doi),
        ("arxiv_id", lambda e: e.arxiv_id),
        ("normalized_title_year", lambda e: f"{normalize_title(e.title)}::{e.year}" if e.title and e.year else ""),
    ):
        bucket: Dict[str, List[BibEntry]] = {}
        for entry in entries:
            value = key_fn(entry)
            if not value:
                continue
            bucket.setdefault(value, []).append(entry)
        for value, dup_entries in bucket.items():
            if len(dup_entries) < 2:
                continue
            groups.append(
                DuplicateGroup(
                    duplicate_type=duplicate_type,
                    value=value,
                    keys=[e.key for e in dup_entries],
                    titles=[e.title for e in dup_entries],
                )
            )
    groups.sort(key=lambda x: (x.duplicate_type, x.value))
    return groups


def venue_field_name(entry: BibEntry) -> str:
    if "journal" in entry.fields:
        return "journal"
    if "booktitle" in entry.fields:
        return "booktitle"
    if entry.entry_type in {"inproceedings", "conference", "proceedings"}:
        return "booktitle"
    return "journal"


def issue_to_field_names(entry: BibEntry, issues: str) -> List[str]:
    field_names: List[str] = []
    for issue in [x for x in (issues or "").split(";") if x]:
        if issue == "title_mismatch":
            field_names.append("title")
        elif issue == "author_mismatch":
            field_names.append("author")
        elif issue == "year_mismatch":
            field_names.append("year")
        elif issue == "venue_mismatch":
            field_names.append(venue_field_name(entry))
        elif issue == "doi_mismatch":
            field_names.append("doi")
        elif issue == "arxiv_mismatch":
            field_names.append("eprint")
    return list(dict.fromkeys(field_names))


def highlight_bib_fields(raw_bib: str, field_names: List[str]) -> str:
    if not raw_bib or not field_names:
        return raw_bib
    highlighted = raw_bib
    for field_name in field_names:
        pattern = re.compile(rf"(^\s*{re.escape(field_name)}\s*=.*(?:\n(?!\s*[A-Za-z][A-Za-z0-9_-]*\s*=).*)*)", re.MULTILINE)
        highlighted = pattern.sub(r"[[[HIGHLIGHT]]]\1[[[/HIGHLIGHT]]]", highlighted)
    return highlighted


def strip_highlight_tokens(value: str) -> str:
    return (value or "").replace("[[[HIGHLIGHT]]]", "").replace("[[[/HIGHLIGHT]]]", "")


def build_recommended_bib(entry: BibEntry, result: CheckResult) -> str:
    fields = dict(entry.fields)
    if result.matched_title:
        fields["title"] = result.matched_title
    if result.matched_authors:
        fields["author"] = " and ".join([x.strip() for x in result.matched_authors.split(";") if x.strip()])
    if result.matched_year:
        fields["year"] = str(result.matched_year)
    if result.matched_venue:
        fields[venue_field_name(entry)] = result.matched_venue
    if result.matched_doi:
        fields["doi"] = result.matched_doi
    if result.matched_url:
        fields["url"] = result.matched_url

    preferred_order = [
        "title",
        "author",
        venue_field_name(entry),
        "year",
        "doi",
        "url",
        "publisher",
        "volume",
        "number",
        "pages",
        "eprint",
    ]
    lines = [f"@{entry.entry_type}{{{entry.key},"]
    seen = set()
    for name in preferred_order:
        value = (fields.get(name) or "").strip()
        if not value:
            continue
        seen.add(name)
        lines.append(f"  {name}={{" + value + "},")
    for name, value in fields.items():
        value = (value or "").strip()
        if not value or name in seen:
            continue
        lines.append(f"  {name}={{" + value + "},")
    if len(lines) > 1:
        lines[-1] = lines[-1].rstrip(",")
    lines.append("}")
    return "\n".join(lines)


def replace_bib_entry_text(full_text: str, target_key: str, replacement_raw: str) -> str:
    pos = 0
    output: List[str] = []
    replaced = False
    for _entry_type, key, raw in iter_bib_entries(full_text):
        start = full_text.find(raw, pos)
        if start < 0:
            continue
        output.append(full_text[pos:start])
        if key == target_key and not replaced:
            output.append(replacement_raw.rstrip() + "\n")
            replaced = True
        else:
            output.append(raw)
        pos = start + len(raw)
    output.append(full_text[pos:])
    if not replaced:
        raise KeyError(target_key)
    return "".join(output)


def build_mismatch_payload(entry: BibEntry, result: CheckResult) -> Dict[str, Any]:
    mismatch_fields = issue_to_field_names(entry, result.issues)
    recommendation_url = result.matched_url or ""
    if not recommendation_url and result.matched_doi:
        recommendation_url = f"https://doi.org/{result.matched_doi}"
    evidence = {
        "source": result.matched_source,
        "title": result.matched_title,
        "authors": [x.strip() for x in result.matched_authors.split(";") if x.strip()],
        "year": result.matched_year,
        "venue": result.matched_venue,
        "doi": result.matched_doi,
        "url": result.matched_url,
    }
    return {
        "mismatch_fields": mismatch_fields,
        "highlighted_raw_bib": highlight_bib_fields(entry.raw, mismatch_fields),
        "recommended_bib": build_recommended_bib(entry, result),
        "evidence": evidence,
        "recommendation_url": recommendation_url,
    }


def apply_bib_correction(job_store: JobStore, job_id: str, key: str, new_bib_text: str) -> None:
    current_bib = job_store.read_modified_bib(job_id)
    updated = replace_bib_entry_text(current_bib, key, strip_highlight_tokens(new_bib_text).strip())
    job_store.write_modified_bib(job_id, updated)
    job = job_store.get(job_id)
    if job and job.output:
        for entry in job.output.entries:
            if entry.get("key") == key:
                entry["raw_bib"] = strip_highlight_tokens(new_bib_text).strip()
                if entry.get("mismatch_details"):
                    entry["mismatch_details"]["recommended_bib"] = strip_highlight_tokens(new_bib_text).strip()
        job_store.update(job_id, output=job.output)


def make_namespace(config: AnalysisConfig, source: SourceOption) -> argparse.Namespace:
    return argparse.Namespace(
        backend=source.name,
        cache=config.cache_path,
        serpapi_key=source.api_key,
        proxy=source.proxy,
        delay=config.delay,
        jitter=config.jitter,
        timeout=config.timeout,
        retries=config.retries,
        retry_backoff=config.retry_backoff,
        num_results=config.num_results,
        min_sources=1,
        year_window=config.year_window,
        title_threshold=config.title_threshold,
        loose_title_threshold=config.loose_title_threshold,
        venue_threshold=config.venue_threshold,
        min_author_overlap=config.min_author_overlap,
        author_check_count=config.author_check_count,
        no_author_in_query=config.no_author_in_query,
        skip_url_check=config.skip_url_check,
        include_missing_title=config.include_missing_title,
        user_agent=config.user_agent,
        verbose=config.verbose,
        dry_run=False,
    )


def verdict_priority(verdict: str) -> int:
    return VERDICT_PRIORITY.get(verdict, 0)


def normalize_backend_verdict(result: CheckResult) -> CheckResult:
    if result.verdict == "ERROR":
        result.verdict = "NOT_FOUND"
        result.exists = False
        issues = [x for x in (result.issues or "").split(";") if x]
        issues.append("source_error")
        result.issues = ";".join(dict.fromkeys(issues))
    return result


def check_entry_with_sources(
    entry: BibEntry,
    config: AnalysisConfig,
    searchers: Dict[str, ReferenceSearcher],
    url_searcher: ReferenceSearcher,
) -> Dict[str, Any]:
    kind = record_kind(entry)
    category = reference_category(entry)
    source_errors: List[Dict[str, str]] = []
    url_ok = None if config.skip_url_check else url_searcher.check_url(entry.url)

    if kind in {"github", "blog"}:
        args = make_namespace(config, SourceOption(name="url"))
        result = decide(entry, [], "web", args, url_ok)
        payload = asdict(result)
        payload["raw_bib"] = entry.raw
        payload["input_authors_list"] = entry.authors
        payload["mismatch_details"] = None
        payload["source_errors"] = source_errors
        payload["reference_category"] = category
        return payload

    best_result: Optional[CheckResult] = None
    for source in config.source_order:
        if not source.enabled:
            continue
        if source.name == "serpapi" and not source.api_key:
            continue
        searcher = searchers.get(source.name)
        if not searcher:
            continue
        args = make_namespace(config, source)
        try:
            candidates = searcher.search(entry)
            result = decide(entry, candidates, source.name, args, url_ok)
            result = normalize_backend_verdict(result)
        except Exception as exc:
            source_errors.append({"source": source.name, "error": str(exc)})
            continue
        if best_result is None or verdict_priority(result.verdict) > verdict_priority(best_result.verdict):
            best_result = result
        elif best_result and verdict_priority(result.verdict) == verdict_priority(best_result.verdict):
            if result.confidence > best_result.confidence:
                best_result = result
        if best_result and best_result.verdict == "FOUND_MATCH":
            break

    if best_result is None:
        fallback = make_namespace(config, SourceOption(name="none"))
        best_result = normalize_backend_verdict(
            decide(entry, [], "none", fallback, url_ok)
        )

    payload = asdict(best_result)
    payload["raw_bib"] = entry.raw
    payload["input_authors_list"] = entry.authors
    payload["reference_category"] = category
    payload["mismatch_details"] = (
        build_mismatch_payload(entry, best_result)
        if best_result.verdict == "FOUND_MISMATCH"
        else None
    )
    payload["source_errors"] = source_errors
    return payload


def analyze_bib_text(text: str, config: AnalysisConfig) -> AnalysisOutput:
    return analyze_bib_text_with_progress(text, config)


def analyze_bib_text_with_progress(
    text: str,
    config: AnalysisConfig,
    progress_callback: Optional[Any] = None,
) -> AnalysisOutput:
    entries = load_entries_from_bib_text(text, include_missing_title=config.include_missing_title)
    duplicates = detect_duplicates(entries)
    searchers: Dict[str, ReferenceSearcher] = {}
    url_searcher = ReferenceSearcher(make_namespace(config, SourceOption(name="arxiv")))
    for source in config.source_order:
        if not source.enabled:
            continue
        if source.name == "serpapi" and not source.api_key:
            continue
        searchers[source.name] = ReferenceSearcher(make_namespace(config, source))

    results: List[Dict[str, Any]] = []
    if progress_callback:
        progress_callback(0, len(entries), "Preparing verification...")
    for idx, entry in enumerate(entries):
        results.append(check_entry_with_sources(entry, config, searchers, url_searcher))
        if progress_callback:
            progress_callback(idx + 1, len(entries), f"Verifying {idx + 1}/{len(entries)}: {entry.key}")
        if idx < len(entries) - 1:
            time.sleep(max(0.0, config.delay + random.random() * config.jitter))

    verdict_counts: Dict[str, int] = {}
    for row in results:
        verdict_counts[row["verdict"]] = verdict_counts.get(row["verdict"], 0) + 1

    duplicate_entry_count = sum(len(group.keys) for group in duplicates)
    duplicate_keys = {key for group in duplicates for key in group.keys}
    summary = AnalysisSummary(
        total_entries=len(entries),
        unique_entries=len(entries) - len(duplicate_keys) + len(duplicates),
        duplicate_groups=len(duplicates),
        duplicate_entries=duplicate_entry_count,
        verdict_counts=verdict_counts,
    )
    return AnalysisOutput(summary=summary, duplicates=duplicates, entries=results)


def run_job(job_store: JobStore, job_id: str, text: str) -> None:
    job = job_store.get(job_id)
    if not job:
        return
    job_store.update(
        job_id,
        status="running",
        error="",
        total_entries=0,
        processed_entries=0,
        progress=0.0,
        progress_message="Preparing verification...",
    )
    try:
        def progress_callback(processed: int, total: int, message: str) -> None:
            progress = 0.0 if total <= 0 else round(processed / total, 4)
            job_store.update(
                job_id,
                total_entries=total,
                processed_entries=processed,
                progress=progress,
                progress_message=message,
            )

        output = analyze_bib_text_with_progress(text, job.config, progress_callback=progress_callback)
        job_store.update(
            job_id,
            status="completed",
            output=output,
            processed_entries=output.summary.total_entries,
            total_entries=output.summary.total_entries,
            progress=1.0,
            progress_message="Verification completed.",
        )
    except Exception as exc:
        job_store.update(job_id, status="failed", error=str(exc), progress_message="Verification failed.")
